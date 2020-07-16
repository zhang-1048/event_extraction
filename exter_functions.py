#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2020-07-10 17:31
# @Author : XXX
# @Site : 
# @File : exter_functions.py
# @Software: PyCharm
import numpy as np
from gensim.models import word2vec
from pyltp import Postagger
import jieba
import openpyxl
import time

postagger = Postagger()
postagger.load("E:\Data\ltp_data_v3.4.0\pos.model")
model = word2vec.Word2Vec.load('../vector_model/all_data_vector')


def get_first_filter_process(data, kws, specific_words):
    """
    data:已经抽取出来的数据
    kws：key_words
    specific_words:一些需要召回的特定的关键词
    """
    original_event = {}
    for i in data:
        for k, v in i.items():
            tmp = []  # 保存当前新闻提取出的  事件短语
            for item in v:
                mm = "".join(item[:3])
                for ikw in kws:
                    if ikw in mm:
                        if len(item) > 3 and len(mm) > 20 and len(item[0]) >= 2 and len(item[2]) >= 3 and len(
                                item[1]) > 1:
                            if len(tmp) == 0: tmp.append(item[:3])
                            c_tmp = 0
                            for itmp in tmp:
                                if mm == "".join(itmp):  # 避免一篇文章中，出现句相同的事件描述句子
                                    c_tmp = 1
                                    # print("相似：", mm,"***", itmp)
                                    break
                            if c_tmp == 0:
                                tmp.append(item[:3])
                        else:
                            if 12 < len(mm) <= 20:
                                for word in specific_words:
                                    if word in mm and item[:3] not in tmp:
                                        tmp.append(item[:3])
                                        break
                        break
            original_event[k] = tmp
    return original_event


def extraction_function(triple_extractor, k_day, v_content, restore_extraction_file, total_count):
    """
    triple_extraction:用于抽取事件的工具
    key_day:抽取事件以每天为单位进行，每一个key_day表示一天
    v_content：表示每个key_day中的全部新闻集合
    restore_extraction_file：抽取结果的宝保存路径
    total_count：新闻总数，用于显示目前的抽取记录
    """
    tmp_sum = 0
    time1 = time.time()
    print(time.asctime(time.localtime(time1)))
    part_list = []
    c = 1
    for index in range(104, len(k_day)):
        print(k_day[index])  # 每个时间区间的起点
        k_list = {}  # 保存当前时间区间内提取出的全部  事件短语

        i_link, i_content = list(v_content[index].keys()), list(v_content[index].values())
        for i_lin, i_con in zip(i_link, i_content):
            svos = triple_extractor.triples_main(i_con)
            tmp_sum += len(svos)
            print("processing:", i_lin, c, "/", total_count, len(svos), tmp_sum)
            c += 1

            k_list[i_lin] = svos
        part_list.append(k_list)
    time2 = time.time()
    print(time2 - time1)
    print(time.asctime(time.localtime(time2)))
    print("保存数据")
    np.save(restore_extraction_file, part_list)  # 104,261


def replace_pure_num(x):
    """
    将x中的数字，替换为NUM，并计算该句子的嵌入向量
    """
    x = str(x).replace("\u3000", "")
    cut_list = list(jieba.lcut(x))
    for index in range(len(cut_list)):
        if str.isdigit(cut_list[index]):
            cut_list[index] = "NUM"

    v = []
    for k in cut_list:
        try:
            v.append(model[k].tolist())
        except KeyError:
            v.append([0.5 for _ in range(100)])
    v1 = np.sum(v, axis=0)
    return v1


def get_most_pure_event_nearest(dict_event_index_link, simi_metrics):
    """
    dict_event_index_link:源链接，形式：  {1：url1，2：url2}
    simi_metrics:事件句子之间的相似度矩阵
    return:Dict,聚类结果，形式：{1：[1，2，3，7，8]，2：[5，6，10]}
    """
    # 计算每个类别，每个事件句与其余全部事件句的相似度，并按照最相似的归类
    c = 1
    keys = dict_event_index_link.keys()
    res_dict = {"1": [0]}  # 表示每个事件对应的事件句的链接和正文

    for index in range(1, len(keys)):  # 表示从values的位置1开始依次遍历
        min_simi = 0.83
        tmp_keys = res_dict.keys()
        index_1 = -1  # 用于记录和当前的index相似度最小的类簇，最小值中的最大值
        max_simi = 0.9  # 最大相似度的起始值
        index_2 = -1  # 用于记录和当前的index相似度最大的类簇
        for k1 in tmp_keys:  # k1表示具体到某个事件的下标

            v1 = res_dict[k1]  # 事件序号为K1的，每一条事件句的下标集合
            # k2:具体到每个事件的下标
            simi_tmp = [simi_metrics[index][k2] for k2 in v1]

            if np.max(simi_tmp) >= max_simi:
                max_simi = np.max(simi_tmp)
                index_2 = k1
                continue
            if np.min(simi_tmp) > min_simi:
                index_1 = k1
                min_simi = np.min(simi_tmp)
        if index_2 != -1:  # 表示该index已经被归类，不需要进行后续
            res_dict[index_2].append(index)
            continue

        if index_1 == -1:  # 表示该index，与任何的簇类的相似度，在0.85至0.9之间
            res_dict[len(tmp_keys) + 1] = [index]
        else:  # 表示该index，与任何的簇类的相似度，军小于0.85
            res_dict[index_1].append(index)

        c += 1

    return res_dict


def cosine_similarity(x, y, norm=False):
    """ 计算两个向量x和y的余弦相似度 """

    res = np.array([[x[i] * y[i], x[i] * x[i], y[i] * y[i]] for i in range(len(x))])
    cos = sum(res[:, 0]) / (np.sqrt(sum(res[:, 1])) * np.sqrt(sum(res[:, 2])))

    return 0.5 * cos + 0.5 if norm else cos  # 归一化到[0, 1]区间


def match_keyword_cls(strage, enterprise, operators, filter_first):
    """
    strage:List,
    enterprise:List,
    operators:List,
    return：List，
    """
    strage_list = []
    enterprise_list = []
    operators_list = []
    other_list = []
    res_dict = {}
    total = [strage, enterprise, operators]
    c = 0
    for k, v in filter_first.items():
        for i in v:
            mark = 0
            for j in strage:
                if j in i:
                    strage_list.append((k, i))
                    mark = 1
                    break
            if mark == 1: continue
            for j in operators:
                if j in i:
                    operators_list.append((k, i))
                    mark = 1
                    break
            if mark == 1: continue
            for j in enterprise:
                if j in i:
                    enterprise_list.append((k, i))
                    mark = 1
                    break
            if mark == 1: continue
            if mark == 0:
                other_list.append((k, i))

    zonghe_list = [strage_list, operators_list, enterprise_list, other_list]
    name_zonghe_list = ["lei1", "lei2", "lei3", "lei4"]

    return zonghe_list, name_zonghe_list


def filter_specific_word(query_sent_list, specific_words, dict_link_title_original, title_allowed):
    """
    specific_words:用于过滤长度在12-20字的特殊句子，将这一部分含有特殊词的句子进行召回
    dict_link_title_original：形式{url1：title1，url2：title2}
    title_allowed:决定是否将标题也加入到相似度的计算中去
    return：filter_first：Dict，词性过滤后的结果，形式：{link:[sentence1,sentence2]}
    """
    filter_first = {}
    posttag_word_list = []
    for k, v in query_sent_list.items():
        #     print(k)
        tmp = []
        for item in v:
            words_split = list(jieba.lcut(item))
            posts = list(postagger.postag(words_split))
            c = 0
            for char in ["ni", "j", "nh", "nz", "ws"]:
                if char in posts:
                    posttag_word_list.append(item)
                    tmp.append(item)
                    c = 1
                    break
            if c == 0:
                for i_word in specific_words:
                    if i_word in words_split:
                        posttag_word_list.append(item)
                        tmp.append(item)
                        break
        if len(tmp) > 0 and title_allowed:
            filter_first[k] = tmp + [dict_link_title_original[k]]  # 将标题也加入，进行聚类

    return filter_first, posttag_word_list


def remove_duplicated_link_in_result(dict_event_index_link, new_combination):
    """
    new_conbination = sorted(zip(v1, add_simi, all_link_simi_with_title, all_link_simi_with_abst),
                         key=lambda x: x[1], reverse=True)
    从最后的结果中，对链接进行去重，避免一个文本中反复出现某个事件，信息出现误判
    """
    res_links, res_list = [], []
    for item in new_combination:
        if dict_event_index_link[item[0]] not in res_links:
            res_links.append(dict_event_index_link[item[0]])
            res_list.append(item)
        else:
            for i in res_list:
                if dict_event_index_link[i[0]] == dict_event_index_link[item[0]]:
                    if i[1] <= item[1]:
                        res_list.remove(i)
                        res_list.append(item)
                        res_links.remove(dict_event_index_link[i[0]])
                        res_links.append(dict_event_index_link[item[0]])
                    break
    return res_list


def event_cluster_function(zonghe_list,
                             name_zonghe_list,
                             dict_link_title_original,
                             dict_link_abstract_infos,
                             dict_link_time,
                             dict_link_contents,
                             dict_link_ids,
                             dict_link_keywords,
                             dict_link_crawl_sites,
                             res_file_name,
                             min_show_num,
                             max_show_num,
                             words_length_limit,
                             remove_duplicated_link):
    """
    zonghe_list：四个分类的数据
    name_zonghe_list：四个分类数据的类别名称
    min_show_num:最小显示数量
    max_show_num:最大显示数量
    res_file_name:保存文件名
    words_length_limit:字数显示限制
    """
    some = "标题信息".split("*")
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    quanju_hang = 1  # 注意：'cell'函数中行列起始值为1
    for ks in range(len(some)):
        outws.cell(column=ks + 1, row=quanju_hang, value=some[ks])
    quanju_hang += 1
    event_counts = 1

    for index in range(len(zonghe_list)):
        dict_name = name_zonghe_list[index]

        print(dict_name)
        tmp_list_tuple = zonghe_list[index]  # 这里不能转为dict，否则同一链接下的语句会被覆盖

        text_num = len(tmp_list_tuple)

        dict_event_index_link = {}
        dict_event_index_event = {}
        data_index = 0
        for k in tmp_list_tuple:
            dict_event_index_link[data_index] = k[0]
            dict_event_index_event[data_index] = k[1].replace("　", "").replace(" ", "")
            data_index += 1

        simi_metrics = np.array([[0.0] * text_num] * text_num, dtype=np.float)
        # 计算相似度矩阵
        for k, v in dict_event_index_event.items():
            for k1, v1 in dict_event_index_event.items():
                if k1 >= k:
                    simi_metrics[k][k1] = cosine_similarity(replace_pure_num(v), replace_pure_num(v1))
                    simi_metrics[k1][k] = simi_metrics[k][k1]
        # 进行聚类
        res1_dict1 = get_most_pure_event_nearest(dict_event_index_link, simi_metrics)
        # 按照事件句的多少进行排序
        res1_dict = dict(sorted(res1_dict1.items(), key=lambda x: len(x[1]), reverse=True))

        outws.cell(column=1, row=quanju_hang, value=dict_name)
        quanju_hang += 1
        # 遍历每一个事件，并进行去重等多个操作
        final_index = 0
        final_dict = {}
        for k1, v1 in res1_dict.items():
            if dict_name in ["lei1", "lei2"] or len(v1) >= min_show_num:  # 对 lei1，或是出现次数大于min_show_num的，才进行展示
                if dict_name != "lei2":  # 进行字数限制，过滤句子太长的结果
                    v1 = [i for i in v1 if len(dict_event_index_event[i]) <= words_length_limit]
                if len(v1) == 0: continue
                # 计算标题相似度
                all_link_simi_with_title = [cosine_similarity(replace_pure_num(dict_event_index_event[k]),
                                                              replace_pure_num(
                                                                  dict_link_title_original[dict_event_index_link[k]]))
                                            for k in v1]
                # 计算摘要相似度
                all_link_simi_with_abst = [cosine_similarity(replace_pure_num(dict_event_index_event[k]),
                                                             replace_pure_num(
                                                                 dict_link_abstract_infos[dict_event_index_link[k]]))
                                           for k in v1]
                # 计算相似度之和
                add_simi = [a + b for a, b in zip(all_link_simi_with_title, all_link_simi_with_abst)]

                # 对数据进行组合，new_combination中每一条表示一条完整的新闻信息,并根据总相似度的大小进行排序
                new_combination = sorted(zip(v1, add_simi, all_link_simi_with_title, all_link_simi_with_abst),
                                         key=lambda x: x[1], reverse=True)

                if remove_duplicated_link:  # 是否选择从结果中删除重复的链接来源文章
                    new_combination = remove_duplicated_link_in_result(dict_event_index_link=dict_event_index_link,
                                                                       new_combination=new_combination)

                # 将每一条结果加入到最后的结果中，并在下面根据事件句的数量进行排序
                final_dict[final_index] = new_combination
                final_index += 1
        # 对最后的句子根据事件句的数量进行排序
        final_dict_sorted = dict(sorted(final_dict.items(), key=lambda x: len(x[1]), reverse=True))

        inner_count = 1
        for even_index, sentences in final_dict_sorted.items():
            if dict_name not in ["lei1", "lei2"] and len(sentences) < min_show_num:
                continue
            t_v, t_add_simi, t_all_link_simi_with_title, t_all_link_simi_with_abst = zip(*sentences)
            F_and_R = ["F"] + ["Re"] * (len(sentences) - 1)
            res_combination = list(zip(F_and_R, t_v, t_add_simi, t_all_link_simi_with_title, t_all_link_simi_with_abst))

            if len(res_combination) > max_show_num:
                res_combination = res_combination[:max_show_num]
            write_res_first = [
                [
                    event_counts,  # 总事件数
                    inner_count,  # 分类内部，事件序号
                    len(list(final_dict_sorted[even_index])),  # 事件句数量
                    c[0],  # F_and_R[
                    dict_event_index_event[c[1]],  # 事件句
                    c[2],  # 总相似度
                    dict_event_index_link[c[1]],  # 来源地址
                    c[3],  # 与标题的相似度
                    dict_link_title_original[dict_event_index_link[c[1]]],  # 标题
                    dict_link_time[dict_event_index_link[c[1]]],  # 发布时间
                    dict_link_contents[dict_event_index_link[c[1]]],  # 正文
                    dict_link_ids[dict_event_index_link[c[1]]],  # id
                    # dict_link_docids[all_link[0]],    # doc_id
                    dict_link_keywords[dict_event_index_link[c[1]]],
                    dict_link_crawl_sites[dict_event_index_link[c[1]]],
                    # dict_link_crawl_keywords[all_link[0]]
                ]
                for c in res_combination  # c=[v1, add_simi, all_link_simi_with_title, all_link_simi_with_abst]
            ]

            for i in range(0, len(write_res_first)):
                for j in range(1, len(write_res_first[0]) + 1):
                    outws.cell(column=j, row=quanju_hang + i, value=write_res_first[i][j - 1])
            quanju_hang += len(write_res_first)
            inner_count += 1
            event_counts += 1
            print("inner conut:", inner_count)
    outwb.save(res_file_name)  # 保存结果
